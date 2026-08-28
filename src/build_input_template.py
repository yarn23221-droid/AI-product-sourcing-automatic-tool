"""
입력 템플릿(data/input/input_template.xlsx) 생성 스크립트.

실행하면 "안내" 시트와 "입력" 시트가 있는 엑셀 파일을 새로 만든다.
- 안내 시트: 이 프로젝트가 무엇인지, 시트를 어떻게 채우는지 설명
- 입력 시트: 실제로 상품 정보를 채워 넣는 곳 (1행 헤더, 2행 설명, 3행부터 데이터)

이 스크립트는 "양식"만 만든다. 실제 계산/분류는 main.py가 다른 모듈을 불러와서 수행한다.
"""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation

# 프로젝트 루트 기준 경로 (src/ 안에서 실행해도, 루트에서 실행해도 동작하도록 절대경로로 계산)
PROJECT_ROOT = Path(__file__).resolve().parent.parent
DEFAULT_OUTPUT_PATH = PROJECT_ROOT / "data" / "input" / "input_template.xlsx"

# 입력 시트에 예시로 미리 채워 넣을 빈 데이터 행 개수 (노란색 필수 입력 배경을 미리 칠해 둘 범위)
EMPTY_ROW_COUNT = 200

# 입력 시트 컬럼 정의: 기획안 4번 "입력 시트 스펙" 표를 그대로 코드로 옮긴 것.
# letter: 엑셀 열 문자, name: 헤더 텍스트, required: 필수 입력 여부(노란색 배경 대상),
# desc: 2행에 들어갈 설명 문구, width: 열 너비
INPUT_COLUMNS = [
    {"letter": "A", "name": "번호", "required": False,
     "desc": "자동 채번 (비워두면 프로그램이 순서대로 번호를 매김)", "width": 6},
    {"letter": "B", "name": "작성일", "required": True,
     "desc": "필수 · YYYY-MM-DD 형식으로 입력", "width": 12},
    {"letter": "C", "name": "참고링크", "required": True,
     "desc": "필수 · 11개 유튜버 채널 중 하나의 영상 링크", "width": 30},
    {"letter": "D", "name": "카테고리", "required": True,
     "desc": "필수 · 디지털가전 하위 카테고리 (예: 무선청소기, 커피머신)", "width": 14},
    {"letter": "E", "name": "브랜드", "required": True,
     "desc": "필수 · 제품 브랜드명", "width": 12},
    {"letter": "F", "name": "제품명", "required": True,
     "desc": "필수 · 제품명", "width": 24},
    {"letter": "G", "name": "제품번호", "required": False,
     "desc": "선택 · 모델명", "width": 14},
    {"letter": "H", "name": "옵션", "required": False,
     "desc": "선택 · 색상/용량 등", "width": 12},
    {"letter": "I", "name": "크로켓 등록여부", "required": True,
     "desc": "필수 · O 또는 X (크로켓 제품 링크 유무로 판단)", "width": 12},
    {"letter": "J", "name": "크로켓 제품 링크", "required": False,
     "desc": "선택 · 등록여부가 O인 경우에만 입력", "width": 30},
    {"letter": "K", "name": "크로켓 가격", "required": False,
     "desc": "선택 · 크로켓 제품 링크를 조사해 실제 판매가(숫자만) 입력", "width": 12},
    {"letter": "L", "name": "채널1명", "required": False,
     "desc": "선택 · 네이버쇼핑 등에서 확인한 판매채널명", "width": 14},
    {"letter": "M", "name": "채널1가", "required": False,
     "desc": "선택 · 채널1 가격(숫자만)", "width": 12},
    {"letter": "N", "name": "채널2명", "required": False,
     "desc": "선택 · 판매채널명", "width": 14},
    {"letter": "O", "name": "채널2가", "required": False,
     "desc": "선택 · 채널2 가격(숫자만)", "width": 12},
    {"letter": "P", "name": "채널3명", "required": False,
     "desc": "선택 · 판매채널명", "width": 14},
    {"letter": "Q", "name": "채널3가", "required": False,
     "desc": "선택 · 채널3 가격(숫자만)", "width": 12},
    {"letter": "R", "name": "수요텍스트파일", "required": False,
     "desc": "선택 · data/demand_texts/ 안의 파일명 (예: sample_01.txt)", "width": 22},
    {"letter": "S", "name": "비고", "required": False,
     "desc": "선택 · 자유 메모", "width": 20},
]

# 기획안 3번에 명시된 상품 후보 소스(유튜버 채널 11개)
CANDIDATE_SOURCE_CHANNELS = [
    "이스모", "아론황", "잇섭", "테크몽", "디에디트", "주연",
    "뻘짓연구소", "엠알", "언더케이지", "귀곰", "눈쟁이",
]

# 기획안 3번에 명시된 직구수요 발생사유 텍스트 소스(커뮤니티 6개)
DEMAND_TEXT_SOURCES = [
    "퀘이사존 (하드웨어/모바일 뉴스 게시판)",
    "루리웹 (PC 뉴스 게시판)",
    "쿨엔조이 (신제품 뉴스 게시판)",
    "보드나라 (신제품 뉴스 게시판)",
    "케이벤치 (PC/모바일/라이프 신제품 뉴스 게시판)",
    "아마존 뉴릴리즈",
]

# 입력 시트에 미리 넣어 두는 예시 행 (실제 데이터 아님, 형식만 보여줌 - 가상 브랜드/가상 제품)
EXAMPLE_ROWS = [
    {
        "작성일": "2026-08-01", "참고링크": "https://youtube.com/watch?v=example1",
        "카테고리": "무선청소기", "브랜드": "노바텍(가상)", "제품명": "노바텍 스틱클린 V3",
        "제품번호": "NVT-V3", "옵션": "그레이", "크로켓 등록여부": "O",
        "크로켓 제품 링크": "https://smartstore.naver.com/example/products/1",
        "크로켓 가격": 289000, "채널1명": "쿠팡", "채널1가": 259000,
        "채널2명": "11번가", "채널2가": 265000, "채널3명": "", "채널3가": "",
        "수요텍스트파일": "sample_01.txt", "비고": "예시 행 - 실제 데이터로 덮어써도 됨",
    },
    {
        "작성일": "2026-08-05", "참고링크": "https://youtube.com/watch?v=example2",
        "카테고리": "커피머신", "브랜드": "브루메이트(가상)", "제품명": "브루메이트 에스프레소 미니",
        "제품번호": "BM-MINI2", "옵션": "화이트", "크로켓 등록여부": "X",
        "크로켓 제품 링크": "", "크로켓 가격": "",
        "채널1명": "아마존", "채널1가": 320000,
        "채널2명": "역직구몰A", "채널2가": 305000,
        "채널3명": "", "채널3가": "",
        "수요텍스트파일": "sample_02.txt", "비고": "예시 행 - 실제 데이터로 덮어써도 됨",
    },
]

HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
DESC_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
DESC_FONT = Font(color="595959", italic=True, size=9)
REQUIRED_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
THIN_BORDER = Border(*(Side(style="thin", color="D9D9D9"),) * 4)


def _write_guide_sheet(ws) -> None:
    """"안내" 시트에 프로젝트 설명과 데이터 수집 기준을 적어 넣는다."""
    ws.column_dimensions["A"].width = 100
    ws.sheet_view.showGridLines = False

    lines = [
        ("AI 기반 해외상품·경쟁가격 분석 자동화 - 입력 가이드", True, 14),
        ("", False, 11),
        ("■ 이 파일은 무엇인가요?", True, 12),
        ("'입력' 시트에 상품 정보를 채워 넣고 src/main.py를 실행하면,", False, 11),
        ("판매채널별 최저가 계산 / 가격차이(%) / 타겟가 제안 / 직구수요 분류까지 자동으로 처리한 결과 엑셀이 생성됩니다.", False, 11),
        ("'상품을 찾는 일'(유튜버·베스트 모니터링)은 사람이 계속 담당하고, 그 뒤의 반복 계산·분류만 자동화하는 것이 목표입니다.", False, 11),
        ("", False, 11),
        ("■ 데이터 입력 원칙 (반드시 지켜주세요)", True, 12),
        ("1. 실제 회사 내부 데이터(매입가, 내부 코멘트 등)는 절대 입력하지 마세요. 이 파일은 포트폴리오/학습용입니다.", False, 11),
        ("2. 카테고리는 '디지털가전'으로 한정합니다. (담배 등 소매가가 법으로 고정된 품목은 대상이 아닙니다)", False, 11),
        ("3. 채널 가격은 사람이 직접 조사해 숫자로 입력합니다. 이 프로그램은 어떤 외부 API도 호출하지 않습니다.", False, 11),
        ("4. 모르는 값은 빈칸으로 두세요. 프로그램은 빈칸을 '확인 필요'/'미입력'으로만 표시하고, 값을 추측해서 채우지 않습니다.", False, 11),
        ("", False, 11),
        ("■ 상품 후보를 어디서 찾나요? (최근 1개월 영상 기준, 국내 정식 출시 제품은 1차 제외)", True, 12),
        (", ".join(CANDIDATE_SOURCE_CHANNELS), False, 11),
        ("", False, 11),
        ("■ 직구수요 발생사유는 어디서 조사하나요? (해당 커뮤니티의 신제품/뉴스 게시판)", True, 12),
        (" / ".join(DEMAND_TEXT_SOURCES), False, 11),
        ("텍스트를 data/demand_texts/ 폴더에 .txt 파일로 저장한 뒤, 입력 시트의 '수요텍스트파일' 열에 파일명을 적으세요.", False, 11),
        ("", False, 11),
        ("■ '입력' 시트 사용법", True, 12),
        ("- 1행: 열 이름 / 2행: 설명(입력하지 않음) / 3행부터: 실제 데이터", False, 11),
        ("- 노란색 배경 칸 = 필수 입력 항목입니다.", False, 11),
        ("- 3~4행에는 형식을 보여주는 예시(가상 상품)가 들어 있습니다. 지우고 실제 데이터로 채워도 됩니다.", False, 11),
        ("", False, 11),
        ("■ 실행 방법", True, 12),
        ("1) 이 파일을 채운 뒤 data/input/input_template.xlsx 이름으로 저장", False, 11),
        ("2) 터미널에서: python src/main.py", False, 11),
        ("3) data/output/result_YYYYMMDD_HHMM.xlsx 파일이 새로 생성됩니다.", False, 11),
    ]

    for row_idx, (text, bold, size) in enumerate(lines, start=1):
        cell = ws.cell(row=row_idx, column=1, value=text)
        cell.font = Font(bold=bold, size=size, color="1F4E78" if bold else "000000")
        cell.alignment = Alignment(wrap_text=True, vertical="top")


def _write_input_sheet(ws) -> None:
    """"입력" 시트에 헤더/설명행/서식/예시행/데이터 검증(O,X 드롭다운)을 채운다."""
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"

    # 1행: 헤더
    for col in INPUT_COLUMNS:
        cell = ws[f"{col['letter']}1"]
        cell.value = col["name"]
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[col["letter"]].width = col["width"]

    # 2행: 설명 (입력하지 않는 행)
    for col in INPUT_COLUMNS:
        cell = ws[f"{col['letter']}2"]
        cell.value = col["desc"]
        cell.fill = DESC_FILL
        cell.font = DESC_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 30

    # 3행부터 EMPTY_ROW_COUNT개 행: 필수 열만 노란 배경을 미리 칠해 둔다 (어디를 채워야 할지 표시)
    data_start_row = 3
    data_end_row = data_start_row + EMPTY_ROW_COUNT - 1
    for col in INPUT_COLUMNS:
        if not col["required"]:
            continue
        for row_idx in range(data_start_row, data_end_row + 1):
            cell = ws[f"{col['letter']}{row_idx}"]
            cell.fill = REQUIRED_FILL
            cell.border = THIN_BORDER

    # 전체 데이터 영역에 옅은 테두리 (필수가 아닌 칸도 표 형태를 갖추도록)
    for col in INPUT_COLUMNS:
        for row_idx in range(data_start_row, data_end_row + 1):
            ws[f"{col['letter']}{row_idx}"].border = THIN_BORDER

    # 크로켓 등록여부(I열) O/X 드롭다운
    dv = DataValidation(type="list", formula1='"O,X"', allow_blank=True, showDropDown=False)
    dv.error = "O 또는 X만 입력할 수 있습니다."
    dv.errorTitle = "잘못된 입력"
    ws.add_data_validation(dv)
    dv.add(f"I{data_start_row}:I{data_end_row}")

    # 예시 행(3~4행) 채우기 - 가상 데이터, 형식 참고용
    name_to_letter = {col["name"]: col["letter"] for col in INPUT_COLUMNS}
    for offset, example in enumerate(EXAMPLE_ROWS):
        row_idx = data_start_row + offset
        ws[f"A{row_idx}"] = offset + 1  # 번호
        for name, value in example.items():
            letter = name_to_letter[name]
            ws[f"{letter}{row_idx}"] = value

    # 자동 필터
    ws.auto_filter.ref = f"A1:{INPUT_COLUMNS[-1]['letter']}{data_end_row}"


def build_input_template(output_path: Path = DEFAULT_OUTPUT_PATH) -> Path:
    """입력 템플릿 엑셀 파일을 생성하고 저장한 뒤 경로를 반환한다."""
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    guide_ws = wb.active
    guide_ws.title = "안내"
    _write_guide_sheet(guide_ws)

    input_ws = wb.create_sheet("입력")
    _write_input_sheet(input_ws)

    wb.save(output_path)
    return output_path


if __name__ == "__main__":
    saved_path = build_input_template()
    print(f"입력 템플릿을 생성했습니다: {saved_path}")
