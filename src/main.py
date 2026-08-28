"""
전체 파이프라인 실행 스크립트.

입력 시트(data/input/input_template.xlsx, "입력" 시트)를 읽어서
- 판매채널별 최저가 계산 (price_analysis.py)
- 직구수요 존재여부/발생사유 분류 (demand_classifier.py)
- 타겟가 제안 (price_analysis.py)
를 거쳐 결과 엑셀(data/output/result_YYYYMMDD_HHMM.xlsx)을 생성한다.

사용법: python src/main.py
"""

import sys
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

from demand_classifier import DEMAND_TEXT_DIR, classify_demand_text
from excel_builder import ResultRow, build_result_excel
from price_analysis import ChannelPrice, find_lowest_channel, suggest_target_price, to_number

PROJECT_ROOT = Path(__file__).resolve().parent.parent
INPUT_PATH = PROJECT_ROOT / "data" / "input" / "input_template.xlsx"
OUTPUT_DIR = PROJECT_ROOT / "data" / "output"
INPUT_SHEET_NAME = "입력"
DATA_START_ROW = 3  # 1행 헤더, 2행 설명, 3행부터 실제 데이터

# 입력 시트의 열 순서 (build_input_template.py의 INPUT_COLUMNS와 같은 순서/의미, A열부터)
INPUT_FIELD_ORDER = [
    "번호", "작성일", "참고링크", "카테고리", "브랜드", "제품명", "제품번호", "옵션",
    "크로켓 등록여부", "크로켓 제품 링크", "크로켓 가격",
    "채널1명", "채널1가", "채널2명", "채널2가", "채널3명", "채널3가",
    "수요텍스트파일", "비고",
]

# 사람이 반드시 채워야 하는 항목 (기획안 4번 표 기준). 비어 있어도 실행은 멈추지 않고 경고만 남긴다.
REQUIRED_FIELDS = ["작성일", "참고링크", "카테고리", "브랜드", "제품명", "크로켓 등록여부"]


def _cell_text(value) -> str:
    """엑셀 셀 값을 화면에 보여줄 문자열로 정리한다 (None -> 빈 문자열, datetime -> YYYY-MM-DD)."""
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    return str(value).strip()


def _read_input_rows(input_path: Path) -> list[dict]:
    """입력 시트를 읽어 실제 데이터가 있는 행만 딕셔너리 목록으로 반환한다.

    브랜드와 제품명이 둘 다 비어 있는 행은 형식만 갖춘 빈 행(템플릿에 미리 칠해 둔 노란 배경 등)으로
    보고 건너뛴다. 사용자가 입력하다 만 실제 데이터를 함부로 지어내지 않기 위해, 그 외 항목이 비어
    있어도 행 자체는 그대로 살려서 결과에 "확인 필요"로 남긴다.
    """
    if not input_path.exists():
        raise FileNotFoundError(
            f"입력 파일을 찾을 수 없습니다: {input_path}\n"
            "먼저 'python src/build_input_template.py'를 실행해 템플릿을 만들고, "
            "상품 정보를 채운 뒤 다시 시도하세요."
        )

    wb = load_workbook(input_path, data_only=True)
    if INPUT_SHEET_NAME not in wb.sheetnames:
        raise ValueError(
            f"'{input_path.name}' 파일 안에 '{INPUT_SHEET_NAME}' 시트가 없습니다. 시트 이름을 확인해주세요."
        )
    ws = wb[INPUT_SHEET_NAME]

    rows = []
    for row_idx in range(DATA_START_ROW, ws.max_row + 1):
        raw_values = [ws.cell(row=row_idx, column=col).value for col in range(1, len(INPUT_FIELD_ORDER) + 1)]
        row_dict = dict(zip(INPUT_FIELD_ORDER, raw_values))

        if _cell_text(row_dict.get("제품명")) == "" and _cell_text(row_dict.get("브랜드")) == "":
            continue  # 실제 데이터가 없는 빈 행은 건너뜀

        rows.append(row_dict)

    return rows


def _process_row(row_number: int, raw_row: dict) -> ResultRow:
    """입력 시트 한 행을 결과 시트 한 행(ResultRow)으로 변환한다."""
    missing_fields = [field for field in REQUIRED_FIELDS if _cell_text(raw_row.get(field)) == ""]

    channels = [
        ChannelPrice(raw_row.get("채널1명"), raw_row.get("채널1가")),
        ChannelPrice(raw_row.get("채널2명"), raw_row.get("채널2가")),
        ChannelPrice(raw_row.get("채널3명"), raw_row.get("채널3가")),
    ]
    lowest_name, lowest_price = find_lowest_channel(channels)

    crocket_price_raw = raw_row.get("크로켓 가격")
    target = suggest_target_price(crocket_price_raw, lowest_price)

    demand = classify_demand_text(raw_row.get("수요텍스트파일"), DEMAND_TEXT_DIR)

    비고 = _cell_text(raw_row.get("비고"))
    if missing_fields:
        note = f"[확인 필요] 필수 항목 미입력: {', '.join(missing_fields)}"
        비고 = f"{비고} {note}".strip() if 비고 else note

    return ResultRow(
        번호=row_number,
        작성일=_cell_text(raw_row.get("작성일")),
        참고링크=_cell_text(raw_row.get("참고링크")),
        카테고리=_cell_text(raw_row.get("카테고리")),
        브랜드=_cell_text(raw_row.get("브랜드")),
        제품명=_cell_text(raw_row.get("제품명")),
        제품번호=_cell_text(raw_row.get("제품번호")),
        옵션=_cell_text(raw_row.get("옵션")),
        직구수요_존재여부=demand.exists,
        직구수요_발생사유=demand.reason,
        크로켓_등록여부=_cell_text(raw_row.get("크로켓 등록여부")),
        크로켓_제품_링크=_cell_text(raw_row.get("크로켓 제품 링크")),
        크로켓_가격=to_number(crocket_price_raw),
        최저가_채널명=lowest_name,
        최저가=lowest_price,
        타겟가=target.target_price,
        타겟가_코멘트=target.comment,
        비고=비고,
    )


def run(input_path: Path = INPUT_PATH, output_dir: Path = OUTPUT_DIR) -> Path:
    """파이프라인 전체를 실행하고 생성된 결과 파일 경로를 반환한다."""
    raw_rows = _read_input_rows(input_path)

    if not raw_rows:
        print(
            "[안내] 입력 시트에 데이터가 없습니다. "
            f"'{input_path}'의 '입력' 시트 3행부터 상품 정보를 채운 뒤 다시 실행해주세요."
        )

    result_rows = []
    for offset, raw_row in enumerate(raw_rows, start=1):
        try:
            result_rows.append(_process_row(offset, raw_row))
        except Exception as error:  # 한 행에서 문제가 생겨도 전체 실행이 멈추지 않도록 함
            product_name = _cell_text(raw_row.get("제품명")) or "(제품명 미입력)"
            print(f"[경고] {offset}번째 행('{product_name}') 처리 중 오류가 발생해 이 행은 건너뜁니다: {error}")

    timestamp = datetime.now().strftime("%Y%m%d_%H%M")
    output_path = output_dir / f"result_{timestamp}.xlsx"
    build_result_excel(result_rows, output_path)

    print(f"[완료] 총 {len(result_rows)}건을 분석해 결과 파일을 생성했습니다: {output_path}")
    return output_path


if __name__ == "__main__":
    try:
        run()
    except FileNotFoundError as error:
        print(f"[오류] {error}")
        sys.exit(1)
    except Exception as error:
        print(f"[오류] 예상치 못한 문제가 발생했습니다: {error}")
        sys.exit(1)
