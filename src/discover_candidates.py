"""
[프로토타입] 네이버 베스트 -> 후보 필터링 -> 크로켓 조회까지 자동으로 묶어서 실행하는 스크립트.

지금은 소규모 검증 단계라 카테고리를 TARGET_SUBCATEGORY_NAMES로 몇 개만 지정해서 돌린다.
실제로 하는 일:
  1. 지정한 하위 카테고리들의 네이버 베스트 일간 랭킹 top N을 가져온다
  2. 카테고리 간에 겹치는 상품은 1번만 남긴다 (productId 기준)
  3. 최근 10일 이내 이미 리스트업했던 상품은 제외한다
  4. 상품명에 국내 대기업 브랜드가 명시적으로 있으면 제외한다 (그 외에는 "해외 후보"로 남김)
  5. 남은 후보들을 크로켓에서 검색해 등록여부/가격(배송비포함)/링크를 확인한다
  6. 결과를 엑셀로 저장한다 (사람이 채워야 하는 항목은 빈칸으로 남김)

사람이 직접 채워야 하는 항목(네이버 가격비교가 봇 차단이라 자동화 불가 - README 2번 참고):
  - 직구수요 존재여부 등급(구매/찜/리뷰수 기반 O/세모/X, 기준은 헤더 셀 메모 참고), 직구수요 발생사유
  - 크로켓 외 판매채널의 2번째/3번째 채널 가격 (채널1은 네이버 베스트 랭킹 자체의 판매처로 자동 입력됨)

브랜드/제품명/제품번호/옵션(사이즈,컬러)은 title_parser.py가 상품명 텍스트에서 규칙 기반으로
뽑아낸 값이라 휴리스틱이다 - 특히 브랜드는 상품명에 브랜드가 안 써 있으면 틀릴 수 있고, 옵션은
제목 끝에 사이즈/색상 단어가 없으면 감지를 못 해 빈칸으로 남으므로 사람이 확인해야 한다.
"""

from datetime import date, datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill

from croket_client import CroketClient
from domestic_brand_filter import is_domestic_brand
from listing_history import filter_out_recent, record_listed
from title_parser import parse_title
import naver_best_client as naver

PROJECT_ROOT = Path(__file__).resolve().parent.parent
OUTPUT_PATH = PROJECT_ROOT / "data" / "output" / f"candidates_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"

# "순수 기기"가 아닌 하위 카테고리는 후보 수집 대상에서 제외한다.
# - 게임기/타이틀: 게임 하드웨어(콘솔/컨트롤러)뿐 아니라 스팀 게임코드 같은 비실물 상품이 섞여 나옴
# - 소프트웨어: 애초에 실물 기기가 아님
NON_DEVICE_SUBCATEGORY_NAMES = {"게임기/타이틀", "소프트웨어"}
TOP_N_PER_CATEGORY = 10  # [시뮬레이션] 넓은 카테고리 커버리지 확인용, 어제 top3는 이미 이력에 있어 10으로 확장
PERIOD_TYPE = "DAILY"

HEADERS = [
    "작성일", "참고링크", "카테고리", "브랜드", "제품명", "제품번호", "옵션",
    "크로켓 등록여부", "크로켓 제품 링크", "크로켓 가격(배송비포함)",
    "채널1명", "채널1가", "채널2명", "채널2가", "채널3명", "채널3가",
    "직구수요 존재여부", "직구수요 발생사유", "비고",
]


def collect_candidates() -> list[dict]:
    """지정된 하위 카테고리들의 베스트 상품을 모아서 카테고리 중복 제거까지 마친 후보 목록을 만든다."""
    all_subs = naver.fetch_subcategories(period_type=PERIOD_TYPE)
    target_subs = [s for s in all_subs if s["categoryName"] not in NON_DEVICE_SUBCATEGORY_NAMES]

    seen_product_ids = set()
    candidates = []
    for sub in target_subs:
        products = naver.fetch_rank(sub["categoryId"], period_type=PERIOD_TYPE, top_n=TOP_N_PER_CATEGORY)
        for p in products:
            if p["productId"] in seen_product_ids:
                continue  # 다른 카테고리에도 겹쳐서 나온 상품 -> 1번만 남김
            seen_product_ids.add(p["productId"])
            p["categoryName"] = sub["categoryName"]
            candidates.append(p)

    return candidates


def filter_candidates(candidates: list[dict]) -> list[dict]:
    """10일 이내 중복 제거 + 국내 대기업 브랜드 제외."""
    product_ids = [c["productId"] for c in candidates]
    kept_ids = set(filter_out_recent(product_ids))

    result = []
    for c in candidates:
        if c["productId"] not in kept_ids:
            continue
        if is_domestic_brand(c["title"]):
            continue
        result.append(c)
    return result


def enrich_with_croket(candidates: list[dict]) -> list[dict]:
    """각 후보를 크로켓에서 검색해 등록여부/가격/링크 정보를 덧붙인다.

    상품 수가 많으면 한두 건에서 타임아웃/일시적 오류가 나는 걸 막을 수 없어서, 한 건이
    실패해도 나머지 진행을 멈추지 않고 그 건만 "확인 필요"로 남긴 채 계속한다.
    """
    with CroketClient() as client:
        for i, c in enumerate(candidates, start=1):
            print(f"      ({i}/{len(candidates)}) {c['title'][:30]}")
            try:
                match = client.find_registered_product(c["title"])
            except Exception as e:
                print(f"      -> 조회 실패, 건너뜀: {e}")
                c["croket_registered"] = "확인 필요"
                c["croket_link"] = ""
                c["croket_price"] = None
                continue
            if match:
                c["croket_registered"] = "O"
                c["croket_link"] = match["linkUrl"]
                c["croket_price"] = match["priceInclShipping"]
            else:
                c["croket_registered"] = "X"
                c["croket_link"] = ""
                c["croket_price"] = None
    return candidates


# 링크 열에 적용할 파란 밑줄 하이퍼링크 스타일
LINK_FONT = Font(color="0563C1", underline="single")

DEMAND_TIER_NOTE = (
    "네이버에서 상품명 검색 -> 가격비교 -> 해외직구 탭에서 확인.\n"
    "구매/찜/리뷰 중 가장 큰 수 기준: 5개 이하=X, 20개 이상=세모, 50개 이상=O"
)


def build_excel(candidates: list[dict], output_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "후보"

    header_fill = PatternFill(start_color="1E5631", end_color="1E5631", fill_type="solid")
    for col_idx, name in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    ws.cell(row=1, column=HEADERS.index("직구수요 존재여부") + 1).comment = Comment(DEMAND_TIER_NOTE, "안내")

    today_str = date.today().strftime("%Y-%m-%d")
    for row_idx, c in enumerate(candidates, start=2):
        delivery_fee = int(c["deliveryFee"]) if str(c["deliveryFee"]).isdigit() else 0
        naver_price_incl_shipping = c["discountPriceValue"] + delivery_fee
        parsed = parse_title(c["title"])

        values = [
            today_str, c["linkUrl"], c["categoryName"],
            parsed["brand"], parsed["product_name"], parsed["model_no"], parsed["option_category"],
            c["croket_registered"], c["croket_link"], c["croket_price"],
            c["mallNm"], naver_price_incl_shipping, "", "", "", "",
            "", "", "",
        ]
        for col_idx, v in enumerate(values, start=1):
            ws.cell(row=row_idx, column=col_idx, value=v)

        link_cell = ws.cell(row=row_idx, column=HEADERS.index("참고링크") + 1)
        link_cell.hyperlink = c["linkUrl"]
        link_cell.font = LINK_FONT
        if c["croket_link"]:
            croket_link_cell = ws.cell(row=row_idx, column=HEADERS.index("크로켓 제품 링크") + 1)
            croket_link_cell.hyperlink = c["croket_link"]
            croket_link_cell.font = LINK_FONT

    widths = [11, 32, 12, 12, 40, 10, 10, 10, 30, 14, 14, 12, 14, 12, 14, 12, 12, 16, 45]
    for col_idx, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{ws.cell(row=1, column=len(HEADERS)).column_letter}{max(len(candidates) + 1, 2)}"

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def run():
    print(f"[1/5] 네이버 베스트에서 후보 수집 중... (순수 기기 카테고리, 카테고리당 top {TOP_N_PER_CATEGORY})")
    candidates = collect_candidates()
    print(f"      -> 카테고리 중복 제거 후 {len(candidates)}건")

    print("[2/5] 10일 이내 중복 / 국내 대기업 브랜드 제외 중...")
    candidates = filter_candidates(candidates)
    print(f"      -> 남은 해외 후보 {len(candidates)}건")

    print("[3/5] 크로켓 등록여부/가격 조회 중... (상품 수에 따라 시간이 걸립니다)")
    candidates = enrich_with_croket(candidates)

    print("[4/5] 결과 엑셀 생성 중...")
    build_excel(candidates, OUTPUT_PATH)

    print("[5/5] 이번에 리스트업한 상품을 이력에 기록 중...")
    record_listed([c["productId"] for c in candidates])

    print(f"\n완료: {OUTPUT_PATH}")
    print(f"총 {len(candidates)}건, 그 중 크로켓 등록됨 {sum(1 for c in candidates if c['croket_registered'] == 'O')}건")


if __name__ == "__main__":
    run()
