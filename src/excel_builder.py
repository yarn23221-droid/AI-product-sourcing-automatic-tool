"""
결과 엑셀(data/output/result_YYYYMMDD_HHMM.xlsx) 생성 모듈.

이 모듈은 이미 계산이 끝난 값들을 받아서 서식 있는 엑셀 파일로 "정리"만 한다.
가격차이(%) 열은 파이썬에서 미리 계산한 숫자를 박아 넣지 않고, 실제 엑셀 수식으로 써 넣는다.
그래야 나중에 크로켓 가격이나 채널 최저가를 엑셀에서 직접 고치면 가격차이(%)도 자동으로 다시 계산된다.
"""

from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Optional

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# 결과 시트에 나올 컬럼 순서 (기획안 6번 스펙 그대로)
RESULT_HEADERS = [
    "번호", "작성일", "참고링크", "카테고리", "브랜드", "제품명", "제품번호", "옵션",
    "직구수요 존재여부", "직구수요 발생사유", "크로켓 등록여부", "크로켓 제품 링크", "크로켓 가격",
    "판매채널(최저가 채널)", "판매채널별 최저가", "가격차이(%)", "타겟가", "비고",
]

# 가격차이(%) 조건부 서식 판단 기준 (price_analysis.py의 기준과 동일하게 맞춘다)
EXPENSIVE_THRESHOLD = 0.05
CHEAP_THRESHOLD = -0.05

HEADER_FILL = PatternFill(start_color="1E5631", end_color="1E5631", fill_type="solid")  # 진한 초록
HEADER_FONT = Font(color="FFFFFF", bold=True)
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

NOT_ENTERED = "미입력"
NEEDS_CHECK = "확인 필요"
PRICE_FORMAT = '#,##0"원"'


@dataclass
class ResultRow:
    """결과 시트 한 행에 들어갈 값들. price_analysis/demand_classifier의 출력을 모아 담는다."""
    번호: int
    작성일: str
    참고링크: str
    카테고리: str
    브랜드: str
    제품명: str
    제품번호: str
    옵션: str
    직구수요_존재여부: str
    직구수요_발생사유: str
    크로켓_등록여부: str
    크로켓_제품_링크: str
    크로켓_가격: Optional[float]
    최저가_채널명: Optional[str]
    최저가: Optional[float]
    타겟가: Optional[float]
    타겟가_코멘트: str
    비고: str


def _write_header(ws) -> None:
    """1행 헤더에 진한 초록 배경 + 흰 글자 서식을 적용한다."""
    for col_idx, name in enumerate(RESULT_HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _write_row(ws, row_idx: int, row: ResultRow) -> None:
    """데이터 한 행을 엑셀에 써 넣는다. 값이 없는 계산 관련 항목은 '미입력'/'확인 필요'로 명시한다."""
    ws.cell(row=row_idx, column=1, value=row.번호)
    ws.cell(row=row_idx, column=2, value=row.작성일)
    ws.cell(row=row_idx, column=3, value=row.참고링크)
    ws.cell(row=row_idx, column=4, value=row.카테고리)
    ws.cell(row=row_idx, column=5, value=row.브랜드)
    ws.cell(row=row_idx, column=6, value=row.제품명)
    ws.cell(row=row_idx, column=7, value=row.제품번호 or "")
    ws.cell(row=row_idx, column=8, value=row.옵션 or "")
    ws.cell(row=row_idx, column=9, value=row.직구수요_존재여부)
    ws.cell(row=row_idx, column=10, value=row.직구수요_발생사유)
    ws.cell(row=row_idx, column=11, value=row.크로켓_등록여부)
    ws.cell(row=row_idx, column=12, value=row.크로켓_제품_링크 or "")

    price_cell = ws.cell(row=row_idx, column=13)
    if row.크로켓_가격 is None:
        price_cell.value = NOT_ENTERED
    else:
        price_cell.value = row.크로켓_가격
        price_cell.number_format = PRICE_FORMAT

    channel_cell = ws.cell(row=row_idx, column=14)
    lowest_cell = ws.cell(row=row_idx, column=15)
    if row.최저가 is None:
        channel_cell.value = NOT_ENTERED
        lowest_cell.value = NOT_ENTERED
    else:
        channel_cell.value = row.최저가_채널명
        lowest_cell.value = row.최저가
        lowest_cell.number_format = PRICE_FORMAT

    # 가격차이(%) : 하드코딩이 아니라 실제 엑셀 수식으로 기입 (M열=크로켓가격, O열=채널최저가)
    diff_cell = ws.cell(row=row_idx, column=16)
    diff_cell.value = (
        f'=IF(OR(NOT(ISNUMBER(M{row_idx})),NOT(ISNUMBER(O{row_idx}))),"",'
        f'(M{row_idx}-O{row_idx})/O{row_idx})'
    )
    diff_cell.number_format = "0.0%"

    target_cell = ws.cell(row=row_idx, column=17)
    if row.타겟가 is None:
        target_cell.value = NEEDS_CHECK
    else:
        target_cell.value = row.타겟가
        target_cell.number_format = PRICE_FORMAT
    if row.타겟가_코멘트:
        # 타겟가 산정 근거를 셀 메모(코멘트)로 남긴다 - 컬럼 스펙을 그대로 지키기 위함
        target_cell.comment = Comment(row.타겟가_코멘트, "price_analysis.py (자동 생성)")

    ws.cell(row=row_idx, column=18, value=row.비고 or "")


def _apply_sheet_formatting(ws, last_row: int) -> None:
    """틀 고정, 자동 필터, 가격차이(%) 조건부 서식, 열 너비를 적용한다."""
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(RESULT_HEADERS))}{last_row}"

    diff_range = f"P2:P{last_row}"
    ws.conditional_formatting.add(
        diff_range,
        CellIsRule(operator="greaterThan", formula=[str(EXPENSIVE_THRESHOLD)], fill=RED_FILL),
    )
    ws.conditional_formatting.add(
        diff_range,
        CellIsRule(operator="lessThan", formula=[str(CHEAP_THRESHOLD)], fill=GREEN_FILL),
    )

    widths = [6, 12, 30, 12, 12, 22, 12, 10, 14, 34, 12, 30, 12, 16, 14, 12, 12, 20]
    for col_idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def _write_summary_sheet(ws, product_count: int, last_row: int) -> None:
    """"요약" 시트: 생성 시각 / 분석 상품 수 / 가격 조정 필요 상품 수(수식) + 한계 안내."""
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 40

    ws["A1"] = "결과 요약"
    ws["A1"].font = Font(bold=True, size=13)

    ws["A3"] = "생성 시각"
    ws["B3"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    ws["A4"] = "분석 상품 수"
    ws["B4"] = product_count

    ws["A5"] = "가격 조정 필요 상품 수"
    if product_count > 0:
        ws["B5"] = f"=COUNTIF('결과'!P2:P{last_row},\">{EXPENSIVE_THRESHOLD}\")"
    else:
        ws["B5"] = 0

    ws["A7"] = "※ 안내"
    ws["A7"].font = Font(bold=True)
    ws["A8"] = (
        "타겟가는 실제 매입 원가 데이터 없이 '채널 최저가 대비 몇 % 수준'으로만 계산한 "
        "휴리스틱(경험적 규칙) 제안입니다. 최종 가격 결정에는 사람의 검수가 반드시 필요합니다."
    )
    ws["A8"].alignment = Alignment(wrap_text=True)
    ws.merge_cells("A8:D8")
    ws.row_dimensions[8].height = 45


def build_result_excel(rows: list[ResultRow], output_path: Path) -> Path:
    """결과 행 목록을 받아 서식이 적용된 엑셀 리포트를 생성하고 저장한다."""
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    result_ws = wb.active
    result_ws.title = "결과"
    _write_header(result_ws)

    last_row = 1
    for offset, row in enumerate(rows, start=2):
        _write_row(result_ws, offset, row)
        last_row = offset

    _apply_sheet_formatting(result_ws, last_row=max(last_row, 2))

    summary_ws = wb.create_sheet("요약")
    _write_summary_sheet(summary_ws, product_count=len(rows), last_row=max(last_row, 2))

    wb.save(output_path)
    return output_path
